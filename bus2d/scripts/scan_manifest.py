#!/usr/bin/env python3
"""
Step 1 — Scan TDSC-ABUS and Duying datasets, build a unified manifest.

Outputs: manifest.json with per-volume, per-lesion entries including
volume path, bbox, Y-extent, BI-RADS grade, and class id.
"""

import argparse
import csv
import hashlib
import json
import os
import sys
import tarfile
from collections import Counter, defaultdict
from pathlib import Path

import nibabel as nib
import numpy as np
import SimpleITK as sitk
import yaml

# ── helpers ──────────────────────────────────────────────────────────

def resolve_env(s, env=None):
    """Expand ${var} references in a string using env dict."""
    if env is None:
        env = {}
    if not isinstance(s, str):
        return s
    while "${" in s:
        start = s.index("${")
        end = s.index("}", start)
        key = s[start + 2 : end]
        val = env.get(key, os.environ.get(key, ""))
        s = s[:start] + str(val) + s[end + 1 :]
    return s


def load_config(path):
    """Load config.yaml and resolve variable references."""
    with open(path) as f:
        cfg = yaml.safe_load(f)
    env = {"data_root": cfg["data_root"], "output_root": cfg["output_root"]}
    # Recursively resolve
    def _resolve(obj):
        if isinstance(obj, str):
            return resolve_env(obj, env)
        if isinstance(obj, dict):
            return {k: _resolve(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [_resolve(v) for v in obj]
        return obj
    return _resolve(cfg)


def birads_to_class_id(birads_str, class_map):
    """Map a BI-RADS grade string to a class id.  Returns None if unknown."""
    grade = str(birads_str).strip().lower()
    for cid, (grade_val, _) in class_map.items():
        if grade_val.lower() == grade:
            return int(cid)
    return None


# ── ABUS scanner ────────────────────────────────────────────────────

def scan_abus(cfg):
    """Scan TDSC-ABUS dataset.  Returns list of volume dicts."""
    abus_cfg = cfg["datasets"]["abus"]
    base = abus_cfg["base"]
    label_map = cfg["abus_label_map"]
    class_map = cfg["classes"]
    volumes = []

    for split in abus_cfg["splits"]:
        split_dir = os.path.join(base, split)
        labels_csv = os.path.join(split_dir, abus_cfg["labels_csv"])
        bbx_csv = os.path.join(split_dir, abus_cfg["bbx_csv"])

        # Read labels
        labels = {}  # case_id -> {label, data_path, mask_path}
        with open(labels_csv) as f:
            for row in csv.DictReader(f):
                cid = int(row["case_id"])
                labels[cid] = {
                    "label": row["label"],
                    "data_path": os.path.join(
                        split_dir, row["data_path"].replace("\\", "/")
                    ),
                    "mask_path": os.path.join(
                        split_dir, row["mask_path"].replace("\\", "/")
                    ),
                }

        # Read bounding boxes
        bboxes = {}  # id -> {cx, cy, cz, lx, ly, lz}
        with open(bbx_csv) as f:
            for row in csv.DictReader(f):
                bid = int(row["id"])
                bboxes[bid] = {
                    "cx": float(row["c_x"]),
                    "cy": float(row["c_y"]),
                    "cz": float(row["c_z"]),
                    "lx": float(row["len_x"]),
                    "ly": float(row["len_y"]),
                    "lz": float(row["len_z"]),
                }

        for cid, info in sorted(labels.items()):
            birads = label_map.get(info["label"])
            if birads is None:
                print(f"  [WARN] ABUS case {cid}: unknown label '{info['label']}', skipping")
                continue
            cls_id = birads_to_class_id(birads, class_map)
            if cls_id is None:
                continue

            bbx = bboxes.get(cid)
            if bbx is None:
                print(f"  [WARN] ABUS case {cid}: no bbox, skipping")
                continue

            # Coronal extent in voxel coords (axis 0 = coronal sweep)
            # bbx x-axis maps to array axis 0 (330 dim, coronal direction)
            y_min = int(bbx["cx"] - bbx["lx"] / 2)
            y_max = int(bbx["cx"] + bbx["lx"] / 2)

            volumes.append({
                "volume_id": f"abus_{cid:03d}",
                "dataset": "abus",
                "data_path": info["data_path"],
                "mask_path": info["mask_path"],
                "label_tar_path": None,
                "spacing_mm": abus_cfg["spacing_mm"],
                "original_split": split.lower(),
                "lesions": [
                    {
                        "lesion_id": 1,
                        "birads": birads,
                        "class_id": cls_id,
                        "bbox_voxel": {
                            "cx": bbx["cx"], "cy": bbx["cy"], "cz": bbx["cz"],
                            "lx": bbx["lx"], "ly": bbx["ly"], "lz": bbx["lz"],
                        },
                        "y_min": y_min,
                        "y_max": y_max,
                    }
                ],
            })

    return volumes


# ── Duying helpers ──────────────────────────────────────────────────

def parse_label_tar(tar_path):
    """Extract annotation JSON from a Label.tar file."""
    with tarfile.open(tar_path, "r") as t:
        for m in t.getmembers():
            if m.name.endswith(".json"):
                return json.load(t.extractfile(m))
    return None


def parse_duying_annotation(json_data):
    """
    Parse Duying Label.tar JSON → list of lesions.

    Each lesion has: lesion_id, coronal_bboxes, y_extent_voxel.

    Coordinate convention from JSON:
    - FileInfo: Width (axis0), Height (axis1, 3mm), Depth (axis2)
    - Spacing: [X_sp, Y_sp, Z_sp]
    - BoundingBoxLabelModel entries:
        SliceType 0 = sagittal (fixed X),  FrameCount = X index
        SliceType 1 = coronal  (fixed Y),  FrameCount = Y index
        SliceType 2 = axial    (fixed Z),  FrameCount = Z index
        p1, p2 in world-mm coords
    """
    file_info = json_data.get("FileInfo", {})
    spacing = file_info.get("Spacing", [1.0, 3.0, 1.0])
    width = file_info.get("Width", 0)    # X dim
    height = file_info.get("Height", 0)  # Y dim (3mm)
    depth = file_info.get("Depth", 0)    # Z dim

    models = json_data.get("Models", {})
    bbox_list = models.get("BoundingBoxLabelModel") or []

    if not bbox_list:
        return [], (width, height, depth), spacing

    # Group bboxes by lesion label
    lesions_raw = defaultdict(list)
    for bb in bbox_list:
        label = bb.get("Label", 1)
        lesions_raw[label].append(bb)

    lesions = []
    for label, bbs in lesions_raw.items():
        coronal_bbs = []
        y_voxels = set()

        for bb in bbs:
            st = bb.get("SliceType")
            fc = bb.get("FrameCount", 0)
            p1 = bb.get("p1", [0, 0, 0])
            p2 = bb.get("p2", [0, 0, 0])

            if st == 1:  # coronal — fixed Y
                # p1, p2 in world mm: X and Z extents
                x_min_px = min(p1[0], p2[0])
                x_max_px = max(p1[0], p2[0])
                z_min_px = min(p1[2], p2[2])
                z_max_px = max(p1[2], p2[2])
                coronal_bbs.append({
                    "y_voxel": fc,
                    "x_min": x_min_px,  # already in 1mm = pixel
                    "x_max": x_max_px,
                    "z_min": z_min_px,
                    "z_max": z_max_px,
                })
                y_voxels.add(fc)

            elif st == 0:  # sagittal — gives Y extent
                y_min_mm = min(p1[1], p2[1])
                y_max_mm = max(p1[1], p2[1])
                y_min_vox = y_min_mm / spacing[1]
                y_max_vox = y_max_mm / spacing[1]
                for yv in range(int(y_min_vox), int(y_max_vox) + 1):
                    y_voxels.add(yv)

            elif st == 2:  # axial — also gives Y extent
                y_min_mm = min(p1[1], p2[1])
                y_max_mm = max(p1[1], p2[1])
                y_min_vox = y_min_mm / spacing[1]
                y_max_vox = y_max_mm / spacing[1]
                for yv in range(int(y_min_vox), int(y_max_vox) + 1):
                    y_voxels.add(yv)

        if not y_voxels:
            continue

        y_min = min(y_voxels)
        y_max = max(y_voxels)

        # If no coronal bbox but we have y-extent, derive bbox from sagittal/axial
        if not coronal_bbs:
            # Try to build a bbox from other views
            x_vals, z_vals = [], []
            for bb in bbs:
                p1, p2 = bb["p1"], bb["p2"]
                st = bb["SliceType"]
                if st == 0:  # sagittal
                    z_vals.extend([p1[2], p2[2]])
                elif st == 2:  # axial
                    x_vals.extend([p1[0], p2[0]])
            if x_vals and z_vals:
                coronal_bbs.append({
                    "y_voxel": (y_min + y_max) // 2,
                    "x_min": min(x_vals),
                    "x_max": max(x_vals),
                    "z_min": min(z_vals),
                    "z_max": max(z_vals),
                })

        if not coronal_bbs:
            continue

        # Use the first coronal bbox as representative for the 2D label
        rep = coronal_bbs[0]
        cx = (rep["x_min"] + rep["x_max"]) / 2
        lx = rep["x_max"] - rep["x_min"]
        cz = (rep["z_min"] + rep["z_max"]) / 2
        lz = rep["z_max"] - rep["z_min"]

        lesions.append({
            "lesion_id": label,
            "y_min": y_min,
            "y_max": y_max,
            "coronal_bbox_px": {"cx": cx, "cz": cz, "lx": lx, "lz": lz},
            "all_coronal_bbs": coronal_bbs,
        })

    return lesions, (width, height, depth), spacing


def parse_birads_excel(excel_path):
    """
    Parse the BI-RADS classification Excel.

    Returns: dict  nii_filename -> list of {lesion_color, birads}
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    result = {}
    current_file = None
    for row in rows[1:]:  # skip header
        _, fname, lesion_color, birads = row[:4] if len(row) >= 4 else (None,) * 4
        if fname is not None:
            current_file = str(fname).strip()
            if current_file not in result:
                result[current_file] = []
        if current_file and lesion_color is not None and birads is not None:
            # Parse lesion_color: "1-红" → label 1, "2-绿" → label 2, etc.
            color_str = str(lesion_color).strip()
            try:
                lesion_label = int(color_str.split("-")[0])
            except (ValueError, IndexError):
                lesion_label = len(result[current_file]) + 1
            result[current_file].append({
                "lesion_label": lesion_label,
                "birads": str(birads).strip().lower(),
            })
    wb.close()
    return result


# ── Duying scanner ──────────────────────────────────────────────────

def find_nii_tar_pairs(directory):
    """
    Recursively find all (.nii, _nii_Label.tar) pairs under a directory.
    Returns list of (nii_path, tar_path).
    """
    pairs = []
    for root, dirs, files in os.walk(directory):
        nii_files = [f for f in files if f.endswith(".nii")]
        for nf in nii_files:
            stem = nf[:-4]  # remove .nii
            tar_name = f"{stem}_nii_Label.tar"
            tar_path = os.path.join(root, tar_name)
            if os.path.exists(tar_path):
                pairs.append((os.path.join(root, nf), tar_path))
    return pairs


def volume_fingerprint(nii_path, max_samples=100000):
    """
    Compute a fingerprint for deduplication: (shape_tuple, mean, std, p5, p95).
    Uses nibabel to read only the header + a sample of data.
    """
    try:
        img = nib.load(nii_path)
        shape = tuple(img.shape)
        data = np.asarray(img.dataobj, dtype=np.float32).ravel()
        # Subsample for speed
        if len(data) > max_samples:
            rng = np.random.RandomState(42)
            idx = rng.choice(len(data), max_samples, replace=False)
            data = data[idx]
        return (
            shape,
            round(float(np.mean(data)), 2),
            round(float(np.std(data)), 2),
            round(float(np.percentile(data, 5)), 2),
            round(float(np.percentile(data, 95)), 2),
        )
    except Exception as e:
        print(f"  [WARN] Failed to fingerprint {nii_path}: {e}")
        return None


def scan_duying(cfg):
    """Scan all Duying sources. Returns list of volume dicts."""
    duying_cfg = cfg["datasets"]["duying"]
    class_map = cfg["classes"]

    # 1. Parse BI-RADS Excel for 已标注 volumes
    excel_path = duying_cfg["sources"]["annotated_batches"]["birads_excel"]
    birads_map = {}
    if os.path.exists(excel_path):
        birads_map = parse_birads_excel(excel_path)
        print(f"  Loaded BI-RADS Excel: {len(birads_map)} files mapped")

    # 2. Collect all NII+tar pairs from all sources
    all_pairs = []  # (nii_path, tar_path, birads_source, default_birads)

    # 2a. BI-RADS class folders (2类, 3类, 4类)
    for folder_cfg in duying_cfg["sources"]["birads_folders"]:
        fpath = folder_cfg["path"]
        default_br = folder_cfg["default_birads"]
        if not os.path.isdir(fpath):
            print(f"  [WARN] Missing folder: {fpath}")
            continue
        pairs = find_nii_tar_pairs(fpath)
        for nii, tar in pairs:
            all_pairs.append((nii, tar, "folder", default_br))
        print(f"  Scanned {fpath}: {len(pairs)} NII+tar pairs")

    # 2b. 已标注 batches
    annot_cfg = duying_cfg["sources"]["annotated_batches"]
    annot_base = annot_cfg["base"]
    for sub in annot_cfg["sub_dirs"]:
        sub_path = os.path.join(annot_base, sub)
        if not os.path.isdir(sub_path):
            continue
        pairs = find_nii_tar_pairs(sub_path)
        for nii, tar in pairs:
            all_pairs.append((nii, tar, "excel", None))
        print(f"  Scanned {sub_path}: {len(pairs)} NII+tar pairs")

    # 2c. ABS batch 3
    abs_cfg = duying_cfg["sources"]["abs_batch3"]
    abs_path = abs_cfg["path"]
    if os.path.isdir(abs_path):
        pairs = find_nii_tar_pairs(abs_path)
        for nii, tar in pairs:
            all_pairs.append((nii, tar, "excel", None))
        print(f"  Scanned {abs_path}: {len(pairs)} NII+tar pairs")

    print(f"  Total Duying NII+tar pairs found: {len(all_pairs)}")

    # 3. Deduplicate by volume fingerprint
    print("  Computing volume fingerprints for deduplication...")
    fp_groups = defaultdict(list)
    for nii, tar, src, dbr in all_pairs:
        fp = volume_fingerprint(nii)
        if fp is not None:
            fp_groups[fp].append((nii, tar, src, dbr))

    deduped = []
    dup_count = 0
    for fp, group in fp_groups.items():
        # Keep the one with the best BI-RADS source: excel > folder
        group.sort(key=lambda x: 0 if x[2] == "excel" else 1)
        deduped.append(group[0])
        if len(group) > 1:
            dup_count += len(group) - 1

    print(f"  Deduplication: {dup_count} duplicates removed, {len(deduped)} unique volumes")

    # 4. Parse annotations and assign BI-RADS
    volumes = []
    skipped_no_annot = 0
    skipped_no_birads = 0

    for nii_path, tar_path, birads_source, default_birads in deduped:
        nii_basename = os.path.basename(nii_path)

        # Parse annotation
        json_data = parse_label_tar(tar_path)
        if json_data is None:
            skipped_no_annot += 1
            continue

        lesions_raw, (w, h, d), spacing = parse_duying_annotation(json_data)
        if not lesions_raw:
            skipped_no_annot += 1
            continue

        # Resolve BI-RADS per lesion
        excel_entries = birads_map.get(nii_basename, [])

        vol_lesions = []
        for les in lesions_raw:
            lid = les["lesion_id"]
            birads = None

            # Try Excel first
            for entry in excel_entries:
                if entry["lesion_label"] == lid:
                    birads = entry["birads"]
                    break

            # Fallback to folder default
            if birads is None and default_birads is not None:
                birads = default_birads

            if birads is None:
                skipped_no_birads += 1
                continue

            cls_id = birads_to_class_id(birads, class_map)
            if cls_id is None:
                skipped_no_birads += 1
                continue

            vol_lesions.append({
                "lesion_id": lid,
                "birads": birads,
                "class_id": cls_id,
                "y_min": les["y_min"],
                "y_max": les["y_max"],
                "coronal_bbox_px": les["coronal_bbox_px"],
            })

        if not vol_lesions:
            continue

        # Generate a stable volume ID from file path
        vol_id = "duying_" + hashlib.md5(nii_path.encode()).hexdigest()[:8]

        volumes.append({
            "volume_id": vol_id,
            "dataset": "duying",
            "data_path": nii_path,
            "mask_path": None,
            "label_tar_path": tar_path,
            "spacing_mm": spacing,
            "shape": [w, h, d],
            "lesions": vol_lesions,
        })

    print(f"  Duying volumes accepted: {len(volumes)}")
    print(f"  Skipped (no annotation): {skipped_no_annot}")
    print(f"  Skipped lesions (unknown BI-RADS): {skipped_no_birads}")

    return volumes


# ── main ────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Scan datasets and build manifest")
    parser.add_argument(
        "--config",
        default=os.path.join(os.path.dirname(__file__), "..", "config.yaml"),
        help="Path to config.yaml",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print summary only")
    args = parser.parse_args()

    cfg = load_config(args.config)

    print("=" * 60)
    print("BUS-2D Manifest Builder")
    print("=" * 60)

    # Scan ABUS
    print("\n[1/2] Scanning TDSC-ABUS...")
    abus_vols = scan_abus(cfg)
    print(f"  ABUS volumes: {len(abus_vols)}")

    # Scan Duying
    print("\n[2/2] Scanning Duying...")
    duying_vols = scan_duying(cfg)

    all_vols = abus_vols + duying_vols

    # Summary
    print("\n" + "=" * 60)
    print("MANIFEST SUMMARY")
    print("=" * 60)
    print(f"Total volumes: {len(all_vols)}")
    print(f"  ABUS: {len(abus_vols)}")
    print(f"  Duying: {len(duying_vols)}")

    # Class distribution
    class_map = cfg["classes"]
    class_counts = Counter()
    for v in all_vols:
        for les in v["lesions"]:
            class_counts[les["class_id"]] += 1
    print("\nLesion counts by class:")
    for cid in sorted(class_counts):
        grade, name = class_map[cid]
        print(f"  {name} (BI-RADS {grade}): {class_counts[cid]}")
    print(f"  Total lesions: {sum(class_counts.values())}")

    if args.dry_run:
        print("\n[DRY RUN] No files written.")
        return

    # Confirm
    resp = input("\nWrite manifest? [y/N] ").strip().lower()
    if resp != "y":
        print("Aborted.")
        return

    # Write manifest
    out_path = cfg["manifest_path"]
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    manifest = {
        "volumes": all_vols,
        "class_map": {int(k): v for k, v in class_map.items()},
        "total_volumes": len(all_vols),
        "total_lesions": sum(class_counts.values()),
        "class_distribution": {
            class_map[cid][1]: cnt for cid, cnt in class_counts.items()
        },
    }
    with open(out_path, "w") as f:
        json.dump(manifest, f, indent=2, ensure_ascii=False)
    print(f"\nManifest written to: {out_path}")


if __name__ == "__main__":
    main()
